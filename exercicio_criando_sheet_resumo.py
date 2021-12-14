from openpyxl import load_workbook
import os

caminho_arquivo = "C:\\Users\\SylviaMissio\\Desktop\\RPA1\\openpyxl\\Vendedores.xlsx"

plan_aberta = load_workbook(filename=caminho_arquivo)
sheet_seleciona = plan_aberta['Vendas']

somarAmandaMartins = 0
somarElianeMoreira = 0
somarLeonardoAlmeida = 0
somarNicolasPereira = 0

for linha in range(2,len(sheet_seleciona['A']) + 1):
    
    if sheet_seleciona['A%s' % linha].value == "Amanda Martins":
        somarAmandaMartins = somarAmandaMartins + sheet_seleciona['C%s' % linha].value
        
    elif sheet_seleciona['A%s' % linha].value == "Eliane Moreira":
        somarElianeMoreira = somarElianeMoreira + sheet_seleciona['C%s' % linha].value
        
    elif sheet_seleciona['A%s' % linha].value == "Leonardo Almeida":
        somarLeonardoAlmeida = somarLeonardoAlmeida + sheet_seleciona['C%s' % linha].value
    
    elif sheet_seleciona['A%s' % linha].value == "Nicolas Pereira":
        somarNicolasPereira = somarNicolasPereira + sheet_seleciona['C%s' % linha].value

sheet_resumo = plan_aberta.create_sheet(title="Resumo")

sheet_resumo['A1'] = "Vendedores"
sheet_resumo['B1'] = "Vendas"

sheet_resumo['A2'] = "Amanda Martins"
sheet_resumo['B2'] = somarAmandaMartins

sheet_resumo['A3'] = "Eliane Moreira"
sheet_resumo['B3'] = somarElianeMoreira

sheet_resumo['A4'] = "Leonardo Almeida"
sheet_resumo['B4'] = somarLeonardoAlmeida

sheet_resumo['A5'] = "Nicolas Pereira"
sheet_resumo['B5'] = somarNicolasPereira




plan_aberta.save(filename=caminho_arquivo)

os.startfile(caminho_arquivo)